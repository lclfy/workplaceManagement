import openpyxl
import  xlrd
import random
import copy


standard_name = [""]
standard_score = [""]


def get_standard():
    global standard_name
    global standard_score
    workbook = xlrd.open_workbook("standard.xls")
    worksheet = workbook.sheets()[0]
    standard_name = worksheet.col_values(0, start_rowx=1, end_rowx=None)
    standard_score = worksheet.col_values(1, start_rowx=1, end_rowx=None)
    print(standard_name)
    print(standard_score)


def write_workplace_management():
    work_place_workbook = openpyxl.load_workbook("workplace.xlsx")
    all_work_sheets = work_place_workbook.worksheets
    print(random.randint(0, len(standard_name)))

    for ws in all_work_sheets:
        for i in range(4, 22):
            #对于每一行
            #随机找出几个问题和他们的扣分
            wp_problems = ""
            wp_score = 0
            temp_name = standard_name.copy()
            temp_score = standard_score.copy()

            def get_problems():
                nonlocal wp_score
                nonlocal wp_problems
                for count in range(0, random.randint(2, 5)):
                    temp_number = random.randint(0, len(temp_name)-1)
                    wp_problems = wp_problems + "\n" +temp_name[temp_number]
                    wp_score = wp_score + temp_score[temp_number]
                    #删掉已有条目避免重复
                    temp_name.remove(temp_name[temp_number])
                    temp_score.remove(temp_score[temp_number])
                    pass
                #把第一个\n删了
                wp_problems = wp_problems.strip('\n')
                pass
            get_problems()

            ws.cell(i, 3).value = wp_problems
            ws.cell(i, 4).value = wp_score
            ws.cell(i, 5).value = 100-wp_score
            pass
        pass
    work_place_workbook.save('test.xlsx')


if __name__ == '__main__':
    get_standard()
    write_workplace_management()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
